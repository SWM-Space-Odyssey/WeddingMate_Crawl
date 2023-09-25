from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
driver = webdriver.Chrome()
wb = openpyxl.load_workbook('wedding-company.xlsx')
excel_ws = wb['Sheet7']
url = 'https://www.wdgbook.com/partner/hanbok'
category = "한복"
driver.get(url)
count = 1
for i in range(1,8):
    time.sleep(2)
    #페이지 맨 밑으로 내리기
    body = driver.find_element(By.CSS_SELECTOR,'body')
    body.send_keys(Keys.END)
    time.sleep(3)
    for p in range(1,22):
        company_name = driver.find_element(By.XPATH,'//*[@id="main"]/div/section[2]/ul/li['+str(count)+']/a/div[2]/h2').text
        location = driver.find_element(By.XPATH,'//*[@id="main"]/div/section[2]/ul/li['+str(count)+']/a/div[2]/span').text
        
        parts =company_name.split("_")
        result_string = parts[0]    
        
        excel_ws.cell(row=count,column=1).value = result_string
        excel_ws.cell(row=count,column=2).value = location
        excel_ws.cell(row=count,column=3).value = category
        count += 1

        print(result_string)
        print(location)
   
wb.save('wedding-company.xlsx')
